import os
from datetime import date, timedelta
from calendar import monthrange
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from config.database import conectar
from controllers.pago_controller import obtener_estado_venta

COLOR_NAMES = {
    "#FFFFFF": "Blanco",
    "#FACC15": "Amarillo",
    "#EF4444": "Rojo",
    "#3B82F6": "Azul",
    "#22C55E": "Verde",
    "#6B7280": "Gris",
    "#BFC1C2": "Silver Vein",
    "#F97316": "Naranja",
    "#7F1D1D": "Vino",
    "#8B5E3C": "Cafe",
}


def crear_carpeta_exports():
    if not os.path.exists("exports"):
        os.makedirs("exports")


def obtener_rango_semanal():
    hoy = date.today()
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    fin_semana = inicio_semana + timedelta(days=6)

    return inicio_semana.strftime("%Y-%m-%d"), fin_semana.strftime("%Y-%m-%d")


def obtener_rango_mensual():
    hoy = date.today()
    inicio_mes = hoy.replace(day=1)
    ultimo_dia = monthrange(hoy.year, hoy.month)[1]
    fin_mes = hoy.replace(day=ultimo_dia)

    return inicio_mes.strftime("%Y-%m-%d"), fin_mes.strftime("%Y-%m-%d")


def obtener_ventas_por_rango(fecha_inicio, fecha_fin):
    with conectar() as conexion:

        cursor = conexion.execute("""
            SELECT 
                v.id_venta,
                v.fecha,
                c.nombre,
                v.requiere_factura,
                v.total
            FROM ventas v
            INNER JOIN clientes c ON v.id_cliente = c.id_cliente
            WHERE v.fecha BETWEEN ? AND ?
            ORDER BY v.fecha, v.id_venta
        """, (fecha_inicio, fecha_fin))

        ventas = cursor.fetchall()

    resultado = []
    for venta in ventas:
        estado = obtener_estado_venta(venta[0])
        resultado.append({
            "id_venta": venta[0],
            "fecha": venta[1],
            "cliente": venta[2],
            "factura": "Sí" if venta[3] == 1 else "No",
            "total": float(venta[4]),
            "estado": estado
        })

    return resultado


def obtener_pagos_por_rango(fecha_inicio, fecha_fin):
    with conectar() as conexion:

        cursor = conexion.execute("""
            SELECT 
                p.id_pago,
                p.id_venta,
                p.fecha,
                p.monto,
                p.metodo_pago
            FROM pagos p
            WHERE p.fecha BETWEEN ? AND ?
            ORDER BY p.fecha, p.id_pago
        """, (fecha_inicio, fecha_fin))

        pagos = cursor.fetchall()

    return [
        {
            "id_pago": pago[0],
            "id_venta": pago[1],
            "fecha": pago[2],
            "monto": float(pago[3]),
            "metodo_pago": pago[4]
        }
        for pago in pagos
    ]


def obtener_resumen_por_rango(fecha_inicio, fecha_fin):
    ventas = obtener_ventas_por_rango(fecha_inicio, fecha_fin)
    pagos = obtener_pagos_por_rango(fecha_inicio, fecha_fin)

    total_vendido = round(sum(v["total"] for v in ventas), 2)
    total_cobrado = round(sum(p["monto"] for p in pagos), 2)
    saldo_pendiente = round(total_vendido - total_cobrado, 2)
    

    return {
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "total_vendido": total_vendido,
        "total_cobrado": total_cobrado,
        "saldo_pendiente": saldo_pendiente,
        "ventas_pagadas": sum(1 for v in ventas if v["estado"] == "PAGADA"),
        "ventas_abonadas": sum(1 for v in ventas if v["estado"] == "ABONADA"),
        "ventas_pendientes": sum(1 for v in ventas if v["estado"] == "PENDIENTE")
    }


def ajustar_ancho_columnas(ws):
    for columna in ws.columns:
        max_length = 0
        letra_columna = columna[0].column_letter

        for celda in columna:
            if celda.value is not None:
                max_length = max(max_length, len(str(celda.value)))

        ws.column_dimensions[letra_columna].width = max_length + 2


def obtener_color_nombre(color_hex):
    return COLOR_NAMES.get((color_hex or "").upper(), color_hex or "")


def obtener_consulta_producto_por_categoria(categoria, completado=0):
    patron = f"%{categoria.lower()}%"

    with conectar() as conexion:
        cursor = conexion.execute(
            """
            SELECT
                dv.id_detalle,
                c.nombre,
                p.nombre,
                dv.cantidad,
                dv.color,
                dv.leyenda
            FROM detalle_ventas dv
            INNER JOIN ventas v ON dv.id_venta = v.id_venta
            INNER JOIN clientes c ON v.id_cliente = c.id_cliente
            INNER JOIN productos p ON dv.id_producto = p.id_producto
            WHERE LOWER(p.nombre) LIKE ? AND COALESCE(dv.completado, 0) = ?
            ORDER BY c.nombre, p.nombre, dv.id_detalle
            """,
            (patron, 1 if completado else 0),
        )
        filas = cursor.fetchall()

    return [
        {
            "id_detalle": fila[0],
            "cliente": fila[1],
            "producto": fila[2],
            "cantidad": fila[3],
            "color_hex": fila[4],
            "color": obtener_color_nombre(fila[4]),
            "leyenda": fila[5],
        }
        for fila in filas
    ]


def exportar_consulta_producto(categoria, completado=0):
    crear_carpeta_exports()
    registros = obtener_consulta_producto_por_categoria(categoria, completado=completado)

    wb = Workbook()
    ws = wb.active
    sufijo_hoja = "Completados" if completado else "Pendientes"
    ws.title = f"{categoria.capitalize()} {sufijo_hoja}"
    ws.append(["Cliente", "Producto", "Cantidad", "Color", "Leyenda"])

    for registro in registros:
        ws.append([
            registro["cliente"],
            registro["producto"],
            registro["cantidad"],
            registro["color"],
            registro["leyenda"],
        ])
        fila_actual = ws.max_row
        color_hex = (registro["color_hex"] or "#FFFFFF").replace("#", "")
        ws.cell(row=fila_actual, column=4).fill = PatternFill(
            fill_type="solid",
            start_color=color_hex,
            end_color=color_hex,
        )

    ajustar_ancho_columnas(ws)

    sufijo_archivo = "completados" if completado else "pendientes"
    nombre_archivo = f"consulta_{categoria.lower()}_{sufijo_archivo}.xlsx"
    ruta_archivo = os.path.join("exports", nombre_archivo)
    wb.save(ruta_archivo)
    return ruta_archivo


def exportar_reporte_excel(tipo_reporte, fecha_inicio, fecha_fin):
    crear_carpeta_exports()

    resumen = obtener_resumen_por_rango(fecha_inicio, fecha_fin)
    ventas = obtener_ventas_por_rango(fecha_inicio, fecha_fin)
    pagos = obtener_pagos_por_rango(fecha_inicio, fecha_fin)

    wb = Workbook()

    # Hoja resumen
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    ws_resumen.append(["Tipo de reporte", tipo_reporte])
    ws_resumen.append(["Fecha inicio", fecha_inicio])
    ws_resumen.append(["Fecha fin", fecha_fin])
    ws_resumen.append([])
    ws_resumen.append(["Total vendido", resumen["total_vendido"]])
    ws_resumen.append(["Total cobrado", resumen["total_cobrado"]])
    ws_resumen.append(["Saldo pendiente", resumen["saldo_pendiente"]])
    ws_resumen.append(["Ventas pagadas", resumen["ventas_pagadas"]])
    ws_resumen.append(["Ventas abonadas", resumen["ventas_abonadas"]])
    ws_resumen.append(["Ventas pendientes", resumen["ventas_pendientes"]])

    ajustar_ancho_columnas(ws_resumen)

    # Hoja ventas
    ws_ventas = wb.create_sheet("Ventas")
    ws_ventas.append(["ID Venta", "Fecha", "Cliente", "Factura", "Total", "Estado"])

    for venta in ventas:
        ws_ventas.append([
            venta["id_venta"],
            venta["fecha"],
            venta["cliente"],
            venta["factura"],
            venta["total"],
            venta["estado"]
        ])

    ajustar_ancho_columnas(ws_ventas)

    # Hoja pagos
    ws_pagos = wb.create_sheet("Pagos")
    ws_pagos.append(["ID Pago", "ID Venta", "Fecha", "Monto", "Método de pago"])

    for pago in pagos:
        ws_pagos.append([
            pago["id_pago"],
            pago["id_venta"],
            pago["fecha"],
            pago["monto"],
            pago["metodo_pago"]
        ])

    ajustar_ancho_columnas(ws_pagos)

    if tipo_reporte.lower() == "semanal":
        nombre_archivo = f"reporte_semanal_{fecha_inicio}_a_{fecha_fin}.xlsx"
    else:
        nombre_archivo = f"reporte_mensual_{fecha_inicio[:7]}.xlsx"

    ruta_archivo = os.path.join("exports", nombre_archivo)
    wb.save(ruta_archivo)

    return ruta_archivo


def exportar_reporte_semanal():
    fecha_inicio, fecha_fin = obtener_rango_semanal()
    return exportar_reporte_excel("Semanal", fecha_inicio, fecha_fin)


def exportar_reporte_mensual():
    fecha_inicio, fecha_fin = obtener_rango_mensual()
    return exportar_reporte_excel("Mensual", fecha_inicio, fecha_fin)


def exportar_consulta_sillas(completado=0):
    return exportar_consulta_producto("silla", completado=completado)


def exportar_consulta_mesas(completado=0):
    return exportar_consulta_producto("mesa", completado=completado)
